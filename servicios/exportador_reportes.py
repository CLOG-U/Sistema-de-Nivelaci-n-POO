"""Generacion de archivos PDF y Excel para reportes academicos."""

import re
from dataclasses import dataclass
from datetime import date
from io import BytesIO

from interfaz.components.tables import (
    asistencia_registro_to_dict,
    calificacion_registro_to_dict,
    carga_to_dict,
    curso_to_dict,
    usuario_to_dict,
)

TIPOS_REPORTE_ADMIN = [
    "Asistencia",
    "Calificaciones",
    "Inscripciones",
    "Carga academica",
    "General",
]

TIPOS_REPORTE_DOCENTE = [
    "Mis cursos",
    "Calificaciones",
    "Asistencia",
    "Estudiantes inscritos",
]

TIPOS_REPORTE_ESTUDIANTE = [
    "Mis cursos",
    "Mi carga",
    "Mis calificaciones",
    "Mi asistencia",
    "Resumen academico",
]


@dataclass
class ArchivoExportado:
    contenido: bytes
    nombre: str
    mime: str


def _slug(texto: str) -> str:
    limpio = re.sub(r"[^\w\s-]", "", texto, flags=re.UNICODE)
    limpio = re.sub(r"[\s_-]+", "_", limpio.strip())
    return limpio.lower() or "reporte"


def nombre_archivo_reporte(tipo: str, periodo: str, extension: str) -> str:
    return f"reporte_{_slug(tipo)}_{_slug(periodo)}.{extension}"


def tipos_reporte_por_rol(rol: str) -> list[str]:
    if rol == "Administrador":
        return list(TIPOS_REPORTE_ADMIN)
    if rol == "Docente":
        return list(TIPOS_REPORTE_DOCENTE)
    if rol == "Estudiante":
        return list(TIPOS_REPORTE_ESTUDIANTE)
    return []


def _cursos_docente(sistema, docente):
    return [curso for curso in sistema.cursos.values() if curso.docente == docente]


def _filtrar_por_curso(cursos, curso_id):
    if not curso_id or curso_id == "Todos":
        return cursos
    return [curso for curso in cursos if curso.id_curso == curso_id]


def _filas_inscripciones(sistema, periodo, cursos=None):
    filas = []
    for curso in sistema.cursos.values():
        if cursos is not None and curso not in cursos:
            continue
        for estudiante in curso.lista_estudiantes:
            matricula = estudiante.matricula
            if not matricula or matricula.periodo != periodo:
                continue
            filas.append(
                {
                    "Estudiante": f"{estudiante.nombres} {estudiante.apellidos}",
                    "Cedula": estudiante.cedula,
                    "Curso": curso.codigo,
                    "Nombre curso": curso.nombre,
                    "Periodo": matricula.periodo,
                    "Tipo matricula": matricula.tipo_matricula,
                    "Fecha": matricula.fecha_matricula,
                    "Estado": matricula.estado,
                }
            )
    return filas


def obtener_filas_reporte(sistema, tipo_reporte: str, periodo: str) -> list[dict]:
    return obtener_filas_reporte_por_rol(
        sistema,
        "Administrador",
        tipo_reporte,
        periodo,
    )


def obtener_filas_reporte_por_rol(
    sistema,
    rol: str,
    tipo_reporte: str,
    periodo: str,
    usuario=None,
    curso_id=None,
) -> list[dict]:
    if rol == "Administrador":
        return _filas_admin(sistema, tipo_reporte, periodo)

    if rol == "Docente":
        return _filas_docente(sistema, usuario, tipo_reporte, periodo, curso_id)

    if rol == "Estudiante":
        return _filas_estudiante(sistema, usuario, tipo_reporte, periodo)

    return []


def _filas_admin(sistema, tipo_reporte: str, periodo: str) -> list[dict]:
    if tipo_reporte == "Asistencia":
        return [
            asistencia_registro_to_dict(registro)
            for registro in sistema.asistencias.values()
            if registro["periodo"] == periodo
        ]

    if tipo_reporte == "Calificaciones":
        return [
            calificacion_registro_to_dict(registro)
            for registro in sistema.calificaciones.values()
            if registro["periodo"] == periodo
        ]

    if tipo_reporte == "Inscripciones":
        return _filas_inscripciones(sistema, periodo)

    if tipo_reporte == "Carga academica":
        return [
            carga_to_dict(carga)
            for carga in sistema.cargas_academicas.values()
            if carga.periodo == periodo
        ]

    if tipo_reporte == "General":
        resumen = sistema.resumen()
        return [
            {"Indicador": "Usuarios", "Valor": resumen.get("usuarios", 0)},
            {"Indicador": "Docentes", "Valor": resumen.get("docentes", 0)},
            {"Indicador": "Estudiantes", "Valor": resumen.get("estudiantes", 0)},
            {"Indicador": "Cursos", "Valor": resumen.get("cursos", 0)},
            {"Indicador": "Aulas", "Valor": resumen.get("aulas", 0)},
            {"Indicador": "Inscripciones", "Valor": sistema.total_inscripciones()},
            {"Indicador": "Calificaciones", "Valor": resumen.get("calificaciones", 0)},
            {"Indicador": "Asistencias", "Valor": resumen.get("asistencias", 0)},
            {"Indicador": "Cargas academicas", "Valor": resumen.get("cargas", 0)},
            {"Indicador": "Matriculas", "Valor": resumen.get("matriculas", 0)},
            {"Indicador": "Periodo", "Valor": periodo},
        ]

    return []


def _filas_docente(sistema, docente, tipo_reporte: str, periodo: str, curso_id) -> list[dict]:
    if not docente:
        return []

    cursos = _filtrar_por_curso(_cursos_docente(sistema, docente), curso_id)

    if tipo_reporte == "Mis cursos":
        filas = []
        for curso in cursos:
            fila = curso_to_dict(curso)
            fila["Estudiantes inscritos"] = len(curso.lista_estudiantes)
            filas.append(fila)
        return filas

    if tipo_reporte == "Calificaciones":
        return [
            calificacion_registro_to_dict(registro)
            for registro in sistema.calificaciones.values()
            if registro["periodo"] == periodo and registro["curso"] in cursos
        ]

    if tipo_reporte == "Asistencia":
        return [
            asistencia_registro_to_dict(registro)
            for registro in sistema.asistencias.values()
            if registro["periodo"] == periodo and registro["curso"] in cursos
        ]

    if tipo_reporte == "Estudiantes inscritos":
        filas = []
        for curso in cursos:
            for estudiante in curso.lista_estudiantes:
                fila = usuario_to_dict(estudiante)
                fila["Curso"] = curso.codigo
                fila["Nombre curso"] = curso.nombre
                fila["Estado nivelacion"] = estudiante.estado_nivelacion
                filas.append(fila)
        return filas

    return []


def _filas_estudiante(sistema, estudiante, tipo_reporte: str, periodo: str) -> list[dict]:
    if not estudiante:
        return []

    if tipo_reporte == "Mis cursos":
        cursos = sistema.obtener_cursos_estudiante(estudiante)
        return [curso_to_dict(curso) for curso in cursos]

    if tipo_reporte == "Mi carga":
        return [
            carga_to_dict(carga)
            for carga in sistema.cargas_academicas.values()
            if carga.estudiante == estudiante and carga.periodo == periodo
        ]

    if tipo_reporte == "Mis calificaciones":
        return [
            calificacion_registro_to_dict(registro)
            for registro in sistema.obtener_calificaciones_estudiante(estudiante, periodo)
        ]

    if tipo_reporte == "Mi asistencia":
        return [
            asistencia_registro_to_dict(registro)
            for registro in sistema.obtener_asistencias_estudiante(estudiante, periodo)
        ]

    if tipo_reporte == "Resumen academico":
        cursos = sistema.obtener_cursos_estudiante(estudiante)
        calificaciones = sistema.obtener_calificaciones_estudiante(estudiante, periodo)
        asistencias = sistema.obtener_asistencias_estudiante(estudiante, periodo)
        cargas = [
            carga
            for carga in sistema.cargas_academicas.values()
            if carga.estudiante == estudiante and carga.periodo == periodo
        ]
        promedio = "-"
        if calificaciones:
            promedio = round(
                sum(r["calificacion"].nota_final for r in calificaciones) / len(calificaciones),
                2,
            )
        presentes = sum(1 for r in asistencias if r["asistencia"].estado == "Presente")
        return [
            {"Indicador": "Estudiante", "Valor": f"{estudiante.nombres} {estudiante.apellidos}"},
            {"Indicador": "Cedula", "Valor": estudiante.cedula},
            {"Indicador": "Estado nivelacion", "Valor": estudiante.estado_nivelacion},
            {"Indicador": "Periodo", "Valor": periodo},
            {"Indicador": "Cursos inscritos", "Valor": len(cursos)},
            {"Indicador": "Calificaciones", "Valor": len(calificaciones)},
            {"Indicador": "Promedio notas", "Valor": promedio},
            {"Indicador": "Asistencias", "Valor": len(asistencias)},
            {"Indicador": "Asistencias presente", "Valor": presentes},
            {"Indicador": "Cargas academicas", "Valor": len(cargas)},
        ]

    return []


def resumen_filas_reporte(filas: list[dict]) -> str:
    total = len(filas)
    if total == 0:
        return "Sin registros para los filtros seleccionados."
    columnas = list(filas[0].keys())
    return f"{total} registro(s) · columnas: {', '.join(columnas)}"


def generar_excel(datos: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    titulo = Font(bold=True, size=12)
    ws["A1"] = "Sistema de Nivelacion Academica - ULEAM"
    ws["A1"].font = titulo
    ws["A2"] = f"Tipo: {datos.get('tipo', '')}"
    ws["A3"] = f"Periodo: {datos.get('periodo', '')}"
    ws["A4"] = f"Fecha: {datos.get('fecha', '')}"
    ws["A5"] = f"Generado por: {datos.get('generado_por', 'Sistema')}"
    ws["A6"] = f"Descripcion: {datos.get('descripcion', '')}"

    filas = datos.get("filas") or []
    inicio = 8
    if not filas:
        ws[f"A{inicio}"] = "Sin registros para el periodo seleccionado."
    else:
        encabezados = list(filas[0].keys())
        for col, encabezado in enumerate(encabezados, start=1):
            celda = ws.cell(row=inicio, column=col, value=encabezado)
            celda.font = Font(bold=True)
        for offset, fila in enumerate(filas, start=1):
            for col, encabezado in enumerate(encabezados, start=1):
                ws.cell(row=inicio + offset, column=col, value=fila.get(encabezado, ""))

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _texto_seguro(valor) -> str:
    texto = str(valor if valor is not None else "")
    return texto.encode("latin-1", errors="replace").decode("latin-1")


def _render_tabla_pdf(pdf, filas: list[dict], max_columnas: int = 7):
    if not filas:
        pdf.set_font("Helvetica", size=10)
        pdf.cell(0, 8, "Sin registros para los filtros seleccionados.", new_x="LMARGIN", new_y="NEXT")
        return

    encabezados = list(filas[0].keys())[:max_columnas]
    ancho = max(22, (pdf.w - 20) / len(encabezados))

    pdf.set_font("Helvetica", "B", 8)
    for encabezado in encabezados:
        pdf.cell(ancho, 7, _texto_seguro(encabezado)[:16], border=1)
    pdf.ln()

    pdf.set_font("Helvetica", size=7)
    for fila in filas:
        for encabezado in encabezados:
            pdf.cell(ancho, 6, _texto_seguro(fila.get(encabezado, ""))[:20], border=1)
        pdf.ln()


def generar_pdf(datos: dict) -> bytes:
    from fpdf import FPDF

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Sistema de Nivelacion Academica - ULEAM", new_x="LMARGIN", new_y="NEXT", align="C")

    pdf.set_font("Helvetica", size=10)
    pdf.cell(0, 7, f"Tipo de reporte: {_texto_seguro(datos.get('tipo', ''))}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 7, f"Periodo: {_texto_seguro(datos.get('periodo', ''))}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 7, f"Fecha: {_texto_seguro(datos.get('fecha', ''))}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 7, f"Generado por: {_texto_seguro(datos.get('generado_por', 'Sistema'))}", new_x="LMARGIN", new_y="NEXT")
    pdf.multi_cell(0, 7, f"Descripcion: {_texto_seguro(datos.get('descripcion', ''))}")
    pdf.ln(3)

    _render_tabla_pdf(pdf, datos.get("filas") or [])

    salida = pdf.output()
    if isinstance(salida, str):
        return salida.encode("latin-1", errors="replace")
    return bytes(salida)


def exportar_reporte(datos: dict, formato: str) -> ArchivoExportado:
    if formato == "Excel":
        contenido = generar_excel(datos)
        return ArchivoExportado(
            contenido=contenido,
            nombre=nombre_archivo_reporte(datos.get("tipo", "reporte"), datos.get("periodo", ""), "xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    contenido = generar_pdf(datos)
    return ArchivoExportado(
        contenido=contenido,
        nombre=nombre_archivo_reporte(datos.get("tipo", "reporte"), datos.get("periodo", ""), "pdf"),
        mime="application/pdf",
    )


def construir_datos_exportacion(reporte, filas: list[dict], generado_por: str = "Sistema") -> dict:
    return {
        "tipo": reporte.tipo_reporte,
        "periodo": reporte.periodo,
        "descripcion": reporte.descripcion,
        "fecha": reporte.fecha_generacion,
        "generado_por": generado_por,
        "filas": filas,
    }


def preparar_exportacion_directa(
    tipo_reporte: str,
    periodo: str,
    descripcion: str,
    filas: list[dict],
    generado_por: str = "Sistema",
) -> dict:
    return {
        "tipo": tipo_reporte,
        "periodo": periodo,
        "descripcion": descripcion,
        "fecha": date.today().isoformat(),
        "generado_por": generado_por,
        "filas": filas,
    }
